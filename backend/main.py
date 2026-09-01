from __future__ import annotations

import asyncio
import datetime
import io
import json
import os
import re
import secrets
import time
import unicodedata
from typing import Optional

from fastapi import Depends, FastAPI, File, Header, HTTPException, UploadFile, WebSocket, WebSocketDisconnect
from fastapi.responses import FileResponse, Response
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field

import auction as auction_module
import csv_parser
import db
import storage
import push_service

# ========= CONFIG =========
DEFAULT_ADMIN_PASSWORD = "asta2026"
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", DEFAULT_ADMIN_PASSWORD)
if os.environ.get("RAILWAY_ENVIRONMENT_ID") and ADMIN_PASSWORD == DEFAULT_ADMIN_PASSWORD:
    raise RuntimeError("Imposta ADMIN_PASSWORD nelle Variables di Railway prima di avviare Rannatoni.")
FRONTEND_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "frontend"))
DATA_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "data"))
SESSIONS_PATH = os.path.join(DATA_DIR, "sessions.json")
SESSION_SECONDS = 24 * 60 * 60
SIMULATION_PATH = os.path.join(DATA_DIR, "simulation_base.db")

app = FastAPI(title="Rannatoni - Asta Fantacalcio")
db.init_db()
try:
    push_service.ensure_vapid_keys()
except Exception as exc:
    print(f"[push] inizializzazione non riuscita: {exc}")
auction = auction_module.auction
auction.load_persisted()

# ========= SESSIONI =========
def _load_sessions():
    try:
        with open(SESSIONS_PATH, "r", encoding="utf-8") as f:
            raw = json.load(f)
        now = time.time()
        return {k: v for k, v in raw.items() if float(v.get("expires_at", 0)) > now}
    except Exception:
        return {}


def _save_sessions():
    try:
        os.makedirs(DATA_DIR, exist_ok=True)
        with open(SESSIONS_PATH, "w", encoding="utf-8") as f:
            json.dump(SESSIONS, f)
    except Exception as exc:
        print(f"[sessioni] salvataggio non riuscito: {exc}")


SESSIONS: dict[str, dict] = _load_sessions()


def _new_session(role: str, team: str | None):
    token = secrets.token_urlsafe(32)
    now = time.time()
    SESSIONS[token] = {
        "role": role,
        "team": team,
        "created_at": now,
        "expires_at": now + SESSION_SECONDS,
    }
    _save_sessions()
    return token


def _invalidate_team_sessions(team: str, keep_token: str | None = None):
    removed = 0
    for tok, sess in list(SESSIONS.items()):
        if sess.get("role") == "team" and sess.get("team") == team and tok != keep_token:
            SESSIONS.pop(tok, None)
            removed += 1
    if removed:
        _save_sessions()
    return removed


def _token_from_header(authorization: Optional[str]):
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(401, "Non autenticato.")
    return authorization.removeprefix("Bearer ").strip()


def get_session(authorization: Optional[str] = Header(None)):
    token = _token_from_header(authorization)
    session = SESSIONS.get(token)
    if not session or float(session.get("expires_at", 0)) <= time.time():
        SESSIONS.pop(token, None)
        _save_sessions()
        raise HTTPException(401, "Sessione non valida o scaduta.")
    session = dict(session)
    session["token"] = token
    return session


def require_superadmin(session: dict = Depends(get_session)):
    if session.get("role") != "superadmin":
        raise HTTPException(403, "Riservato al Super Admin.")
    return session


def require_team(session: dict = Depends(get_session)):
    if session.get("role") != "team" or not session.get("team"):
        raise HTTPException(403, "Riservato alle squadre.")
    if not db.get_team(session["team"]):
        raise HTTPException(401, "Squadra non piu' valida.")
    return session


def require_spectator(session: dict = Depends(get_session)):
    if session.get("role") != "spectator":
        raise HTTPException(403, "Riservato agli spettatori.")
    return session


def _ensure_pin_changed(session: dict):
    team = db.get_team(session["team"])
    if team and team.get("pin_must_change"):
        raise HTTPException(403, "Prima imposta il tuo PIN personale.")
    return team


def require_auction_manager(session: dict = Depends(require_team)):
    team = _ensure_pin_changed(session)
    if not team or not team.get("is_admin"):
        raise HTTPException(403, "Riservato alla squadra che gestisce l'asta.")
    return session


# ========= PRESENZA / WEBSOCKET =========
class ConnectionManager:
    def __init__(self):
        self.connections: dict[WebSocket, dict] = {}
        self.last_presence: dict[str, dict] = {}
        self.send_locks: dict[WebSocket, asyncio.Lock] = {}

    async def connect(self, ws: WebSocket, session: dict):
        await ws.accept()
        team = session.get("team") if session.get("role") == "team" else None
        self.connections[ws] = {
            "team": team,
            "role": session.get("role"),
            "visibility": "visible",
            "last_seen": time.time(),
        }
        self.send_locks[ws] = asyncio.Lock()
        if team:
            # v10: il login/una sessione attiva equivale a essere entrati nel mercato.
            # Non si torna "non pronti" chiudendo l'app o andando offline: si esce
            # solo con "Ho finito gli acquisti".
            current_team = db.get_team(team)
            if current_team and not current_team.get("market_finished") and not current_team.get("ready"):
                db.set_ready(team, True)
            self.last_presence[team] = {"visibility": "visible", "last_seen": time.time()}

    def touch(self, ws: WebSocket, visibility: str):
        info = self.connections.get(ws)
        if not info:
            return False
        visibility = "hidden" if visibility == "hidden" else "visible"
        changed = visibility != info.get("visibility")
        info["visibility"] = visibility
        info["last_seen"] = time.time()
        if info.get("team"):
            self.last_presence[info["team"]] = {
                "visibility": visibility,
                "last_seen": info["last_seen"],
            }
        return changed

    def disconnect(self, ws: WebSocket):
        info = self.connections.pop(ws, None)
        self.send_locks.pop(ws, None)
        if info and info.get("team"):
            self.last_presence[info["team"]] = {
                "visibility": info.get("visibility", "hidden"),
                "last_seen": time.time(),
            }

    def presence(self, team: str):
        now = time.time()
        active = [i for i in self.connections.values() if i.get("team") == team]
        if any(i.get("visibility") == "visible" and now - i.get("last_seen", 0) < 25 for i in active):
            return "online"
        if any(now - i.get("last_seen", 0) < 90 for i in active):
            return "background"
        last = self.last_presence.get(team)
        if last and now - last.get("last_seen", 0) < 90:
            return "background"
        return "offline"

    def spectator_count(self):
        return sum(1 for info in self.connections.values() if info.get("role") == "spectator")

    async def send_state(self, ws: WebSocket):
        info = self.connections.get(ws) or {}
        team = info.get("team")
        await ws.send_json({"type": "state", "data": build_state(team)})

    async def _send_with_timeout(self, ws: WebSocket, payload: dict, timeout: float = 3.0):
        # Un socket mobile rimasto "zombie" non deve bloccare il realtime degli
        # altri partecipanti. Railway/TCP può impiegare parecchi secondi prima di
        # segnalare una connessione morta: limitiamo ogni invio e trasmettiamo
        # contemporaneamente a tutti i client. Il lock evita due send concorrenti
        # sullo stesso socket (es. broadcast + pong).
        lock = self.send_locks.get(ws)
        if lock is None:
            raise RuntimeError("WebSocket non più connesso")

        async def do_send():
            async with lock:
                await ws.send_json(payload)

        await asyncio.wait_for(do_send(), timeout=timeout)

    async def _broadcast_payloads(self, items):
        async def deliver(ws, payload):
            try:
                await self._send_with_timeout(ws, payload)
                return None
            except Exception:
                return ws

        tasks = [deliver(ws, payload) for ws, payload in items]
        if not tasks:
            return
        dead = await asyncio.gather(*tasks)
        for ws in dead:
            if ws is not None:
                self.disconnect(ws)

    async def broadcast_state(self):
        items = []
        for ws in list(self.connections):
            info = self.connections.get(ws) or {}
            team = info.get("team")
            items.append((ws, {"type": "state", "data": build_state(team)}))
        await self._broadcast_payloads(items)

    async def broadcast_result(self, result: dict):
        await self._broadcast_payloads([
            (ws, {"type": "result", "data": result})
            for ws in list(self.connections)
        ])

    async def broadcast_team_event(self, event_type: str, data: dict):
        """Evento realtime riservato ai Rannatoni autenticati (non spettatori)."""
        await self._broadcast_payloads([
            (ws, {"type": event_type, "data": data})
            for ws, info in list(self.connections.items())
            if info.get("role") == "team"
        ])


manager = ConnectionManager()
AUCTION_LOCK = asyncio.Lock()
_timer_task = None
_presence_task = None
_tocca_task = None
_auto_random_task = None
_auto_random_deadline = None
# Revisione leggera del realtime: i client mobili la controllano ogni secondo
# come rete di sicurezza quando il WebSocket riceve i frame in ritardo.
STATE_VERSION = int(time.time() * 1000)


def _bump_state_version():
    global STATE_VERSION
    STATE_VERSION += 1
    return STATE_VERSION


def _no_store(response: Response):
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"


def simulation_active():
    return db.get_setting("simulation_active", "0") == "1"


def backup_if_real(reason: str):
    if simulation_active():
        return True
    return storage.salva(reason)


def auto_random_enabled():
    return db.get_setting("auto_random", "0") == "1"


def auction_started():
    """Stato ufficiale dell'asta.

    Per database creati prima della v11 inferisce "iniziata" solo se esiste gia'
    almeno un'asta conclusa o una busta attualmente aperta, poi persiste il dato.
    """
    value = db.get_setting("auction_started", None)
    if value is None:
        progress = db.get_auction_progress()
        inferred = auction.mode != "idle" or int(progress.get("auctioned", 0)) > 0
        db.set_setting("auction_started", "1" if inferred else "0")
        return inferred
    return value == "1"


def _ensure_auction_started():
    if not auction_started() and not simulation_active():
        raise HTTPException(400, "L'asta non e' ancora stata avviata dal Super Admin.")


def auto_random_seconds():
    if not _auto_random_deadline:
        return None
    return max(0, int(_auto_random_deadline - time.time() + 0.999))


def _teams_with_presence():
    teams = db.get_public_teams()
    for t in teams:
        t["ready"] = bool(t["ready"])
        t["market_finished"] = bool(t.get("market_finished"))
        t["is_manager"] = bool(t.pop("is_admin"))
        t["presence"] = manager.presence(t["name"])
    return teams


def build_state(viewer_team: str | None = None):
    state = auction.snapshot(viewer_team)
    state["state_version"] = STATE_VERSION
    state["simulation"] = simulation_active()
    state["server_now_ms"] = int(time.time() * 1000)
    state["catalog_revision"] = db.get_setting("catalog_revision", "0")
    state["auction_started"] = auction_started()
    state["auto_random"] = auto_random_enabled()
    state["auto_random_seconds"] = auto_random_seconds()
    state["teams"] = _teams_with_presence()
    team_by_name = {t["name"]: t for t in state["teams"]}
    if state.get("last_result"):
        state["last_result"] = dict(state["last_result"])
        result = state["last_result"]
        player = db.get_player(result["pid"]) if result.get("pid") is not None else None
        if player:
            result["player_out_of_league"] = player["out_of_league"]
        result["released"] = [dict(item) for item in result.get("released", [])]
        for item in result["released"]:
            released_player = db.get_player(item["pid"]) if item.get("pid") is not None else None
            if released_player:
                item["out_of_league"] = released_player["out_of_league"]
        if state["last_result"].get("team"):
            winner = team_by_name.get(state["last_result"]["team"])
            if winner:
                state["last_result"]["team_username"] = winner.get("username") or ""
    state["active_market_count"] = sum(1 for t in state["teams"] if not t.get("market_finished"))
    state["auction_progress"] = db.get_auction_progress()
    rules = db.get_league_settings()
    if auction.mode == "idle":
        state["random_pool"] = {k: v for k, v in db.random_player_pool(rules["random_min_quotation"]).items() if k != "ids"}
    state["league"] = {
        "auction_mode": rules["auction_mode"],
        "max_roster": int(rules["max_roster"]),
        "min_goalkeepers": int(rules["min_goalkeepers"]),
        "max_goalkeepers": int(rules["max_goalkeepers"]),
        "min_outfield": int(rules["min_outfield"]),
        "random_min_quotation": int(rules["random_min_quotation"]),
    }
    if viewer_team:
        me = db.get_team(viewer_team)
        if me:
            roster = db.get_roster(viewer_team)
            state["me"] = {
                "team": viewer_team,
                "username": me.get("username") or "",
                "name_pending": bool(me.get("name_pending")),
                "budget": int(me["budget"]),
                "ready": bool(me["ready"]),
                "is_manager": bool(me["is_admin"]),
                "market_finished": bool(me.get("market_finished")),
                "pin_must_change": bool(me.get("pin_must_change")),
                "roster_size": len(roster),
                "roster_signature": "|".join(f"{g['pid']}:{g['price']}" for g in sorted(roster, key=lambda x: x['pid'])),
                "max_roster": db.max_roster(),
                "roster_rules": db.roster_summary(viewer_team),
            }
            if state["me"]["is_manager"]:
                state["spectator_count"] = manager.spectator_count()
    return state


async def broadcast_state():
    _bump_state_version()
    await manager.broadcast_state()


# ========= AUTH =========
class LoginBody(BaseModel):
    username: str
    pin: str


class AdminLoginBody(BaseModel):
    password: str


@app.post("/api/login")
async def login(body: LoginBody):
    team = db.get_team_by_credentials(body.username.strip(), body.pin.strip())
    if not team:
        raise HTTPException(401, "Username o PIN errati.")
    # v10: il primo login mette automaticamente il Rannatone "in asta".
    # Chi ha gia' concluso il mercato puo' comunque accedere in sola consultazione.
    if not team.get("market_finished"):
        db.set_ready(team["name"], True)
    token = _new_session("team", team["name"])
    return {
        "token": token,
        "team": team["name"],
        "is_manager": bool(team.get("is_admin")),
        "pin_must_change": bool(team.get("pin_must_change")),
    }


@app.post("/api/admin/login")
def admin_login(body: AdminLoginBody):
    if not secrets.compare_digest(body.password, ADMIN_PASSWORD):
        raise HTTPException(401, "Password admin errata.")
    return {"token": _new_session("superadmin", None)}


@app.post("/api/spectator/login")
def spectator_login():
    return {"token": _new_session("spectator", None)}


@app.post("/api/logout")
async def logout(session: dict = Depends(get_session)):
    # Uscire dalla sessione non rimuove piu' un partecipante dal mercato.
    # L'unico evento che lo esclude dalle aste e' la conclusione del mercato.
    SESSIONS.pop(session["token"], None)
    _save_sessions()
    await broadcast_state()
    if auction.mode == "idle" and auto_random_enabled():
        await _arm_auto_random()
    return {"ok": True}


# ========= SUPER ADMIN: PREPARAZIONE =========
@app.get("/api/admin/stats")
def admin_stats(_: dict = Depends(require_superadmin)):
    return {
        **db.count_rows(),
        "auction_started": auction_started(),
        "league_settings": db.get_league_settings(),
        "stats_labels": db.get_stats_labels(),
        "random_pool": {k: v for k, v in db.random_player_pool().items() if k != "ids"},
    }


@app.post("/api/admin/start-auction")
async def admin_start_auction(_: dict = Depends(require_superadmin)):
    if simulation_active():
        raise HTTPException(400, "Termina la simulazione prima di avviare ufficialmente l'asta.")
    if auction.mode != "idle":
        raise HTTPException(400, "C'e' gia' una busta in corso.")
    if auction_started():
        return {"ok": True, "already_started": True}

    counts = db.count_rows()
    teams = db.get_all_teams(True)
    rules = db.get_league_settings()
    if counts.get("players", 0) <= 0:
        raise HTTPException(400, "Carica prima il Catalogo.")
    if rules["auction_mode"] == "repair":
        if counts.get("roster", 0) <= 0 or not teams:
            raise HTTPException(400, "Carica prima le Rose.")
    else:
        if not teams or len(teams) != int(rules["team_count"]):
            raise HTTPException(400, "Crea prima tutte le squadre previste per l'asta da zero.")
        if counts.get("roster", 0) != 0:
            raise HTTPException(400, "L'asta da zero deve partire con rose vuote.")
        if any(t.get("name_pending") for t in teams):
            raise HTTPException(400, "Prima dell'avvio ogni partecipante deve scegliere il nome della propria squadra.")
    if any(not t.get("username") or not t.get("pin_configured") for t in teams):
        raise HTTPException(400, "Completa prima username e PIN di tutte le squadre.")
    if sum(1 for t in teams if t.get("is_admin")) != 1:
        raise HTTPException(400, "Configura esattamente un Gestore asta.")

    db.mark_auction_session_start()
    db.set_setting("auction_started", "1")
    backup_if_real("avvio ufficiale asta")
    await broadcast_state()
    entered = sum(1 for t in teams if t.get("ready") and not t.get("market_finished"))
    return {"ok": True, "entered": entered, "total": len(teams)}


@app.get("/api/admin/teams")
def admin_teams(_: dict = Depends(require_superadmin)):
    return db.get_all_teams(True)


def _slug(text: str):
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^a-zA-Z0-9]+", "", text).lower()
    return text[:18] or "team"


@app.get("/api/admin/teams/suggestions")
def team_suggestions(_: dict = Depends(require_superadmin)):
    teams = db.get_all_teams(True)
    used = set()
    out = []
    for t in teams:
        username = (t.get("username") or "").strip()
        if not username:
            base = _slug(t["name"])
            username = base
            n = 2
            while username.lower() in used:
                username = f"{base}{n}"
                n += 1
        used.add(username.lower())
        # Il Super Admin non puo' leggere un PIN gia' impostato. Per gli account
        # nuovi generiamo solo un PIN temporaneo da distribuire una volta.
        pin = None if t.get("pin_configured") else str(secrets.randbelow(9000) + 1000)
        out.append({
            "name": t["name"], "username": username, "pin": pin,
            "pin_configured": bool(t.get("pin_configured")),
            "budget": int(t["budget"]), "is_admin": bool(t["is_admin"]),
        })
    return out


class TeamConfigItem(BaseModel):
    name: str
    original_name: Optional[str] = None
    username: str
    pin: Optional[str] = None
    budget: int
    is_admin: bool = False


class TeamConfigBody(BaseModel):
    teams: list[TeamConfigItem]


class ZeroTeamCreateBody(BaseModel):
    username: str
    pin: str
    name: Optional[str] = None


class ZeroTeamUpdateBody(BaseModel):
    username: str
    name: Optional[str] = None


class ZeroManagerBody(BaseModel):
    uid: str


class LeagueSettingsBody(BaseModel):
    auction_mode: str = "repair"
    team_count: int = 12
    initial_budget: int = 500
    max_roster: int = 35
    min_goalkeepers: int = 2
    max_goalkeepers: int = 5
    min_outfield: int = 21
    bid_duration: int = 60
    auto_random_delay: int = 10
    random_min_quotation: int = Field(default=0, ge=0)


@app.post("/api/admin/league-settings")
async def save_league_settings(body: LeagueSettingsBody, _: dict = Depends(require_superadmin)):
    if simulation_active():
        raise HTTPException(400, "Termina la simulazione prima di modificare le impostazioni della lega.")
    async with AUCTION_LOCK:
        try:
            values = body.model_dump()
            if auction_started() or auction.mode != "idle":
                current = db.get_league_settings()
                cfg = db.validate_league_settings(values)
                if any(cfg[k] != current[k] for k in cfg if k != "random_min_quotation"):
                    raise ValueError("Dopo l'avvio puoi modificare solo la quotazione minima Mantra per RANDOM.")
                # Il filtro riguarda la prossima estrazione: non tocca rose,
                # crediti, partecipanti o la busta eventualmente gia' aperta.
                db.set_setting("league_random_min_quotation", str(cfg["random_min_quotation"]))
                result = {"settings": cfg, "setup_rebuilt": False}
            else:
                result = db.apply_league_settings(values)
        except ValueError as exc:
            raise HTTPException(400, str(exc)) from exc
    if result.get("setup_rebuilt"):
        _stop_timer()
        _cancel_tocca_task()
        _cancel_auto_random()
        auction.reset(persist=True)
        # Gli slot squadra potrebbero essere stati ricreati: le vecchie sessioni
        # partecipante non devono restare agganciate a nomi non piu' esistenti.
        for tok, info in list(SESSIONS.items()):
            if info.get("role") == "team":
                SESSIONS.pop(tok, None)
        _save_sessions()
        manager.last_presence.clear()
    storage.salva("impostazioni lega")
    if auction.mode == "idle" and auto_random_enabled():
        await _arm_auto_random()
    await broadcast_state()
    return {"ok": True, **result}


@app.post("/api/admin/teams/config")
async def save_team_config(body: TeamConfigBody, _: dict = Depends(require_superadmin)):
    items = [x.model_dump() for x in body.teams]
    try:
        db.save_team_configuration(items)
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    for item in items:
        old_name = str(item.get("original_name") or item.get("name") or "").strip()
        new_name = str(item.get("name") or "").strip()
        if old_name and old_name != new_name:
            _invalidate_team_sessions(old_name)
        if str(item.get("pin") or "").strip():
            _invalidate_team_sessions(new_name)
    storage.salva("configurazione squadre")
    await broadcast_state()
    return {"ok": True}


def _ensure_team_setup_mutable():
    if simulation_active():
        raise HTTPException(400, "Termina la simulazione prima di modificare le squadre.")
    if auction_started() or auction.mode != "idle":
        raise HTTPException(400, "Le squadre si possono modificare solo prima dell'avvio ufficiale.")


@app.post("/api/admin/teams/create")
async def create_zero_team(body: ZeroTeamCreateBody, _: dict = Depends(require_superadmin)):
    _ensure_team_setup_mutable()
    try:
        team = db.create_zero_team(body.username, body.pin, body.name or "")
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    storage.salva("creazione squadra")
    await broadcast_state()
    return {"ok": True, "team": {k: team.get(k) for k in ("uid", "name", "name_pending", "username", "budget")}}


@app.post("/api/admin/teams/{uid}/update")
async def update_zero_team(uid: str, body: ZeroTeamUpdateBody, _: dict = Depends(require_superadmin)):
    _ensure_team_setup_mutable()
    previous = db.get_team_by_uid(uid)
    if not previous:
        raise HTTPException(404, "Squadra non trovata.")
    try:
        team = db.update_zero_team(uid, body.username, body.name)
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    if previous["name"] != team["name"]:
        _invalidate_team_sessions(previous["name"])
    storage.salva("modifica squadra")
    await broadcast_state()
    return {"ok": True}


@app.delete("/api/admin/teams/{uid}")
async def delete_zero_team(uid: str, _: dict = Depends(require_superadmin)):
    _ensure_team_setup_mutable()
    previous = db.get_team_by_uid(uid)
    if not previous:
        raise HTTPException(404, "Squadra non trovata.")
    try:
        db.delete_zero_team(uid)
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    _invalidate_team_sessions(previous["name"])
    storage.salva("eliminazione squadra")
    await broadcast_state()
    return {"ok": True}


@app.post("/api/admin/teams/manager")
async def set_zero_manager(body: ZeroManagerBody, _: dict = Depends(require_superadmin)):
    _ensure_team_setup_mutable()
    if db.auction_mode() != "zero":
        raise HTTPException(400, "Usa la configurazione Rose per scegliere il Gestore nell'asta di riparazione.")
    try:
        db.set_manager_by_uid(body.uid)
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    storage.salva("gestore asta")
    await broadcast_state()
    return {"ok": True}


@app.post("/api/admin/upload/catalog")
async def upload_catalog(file: UploadFile = File(...), _: dict = Depends(require_superadmin)):
    raw = await file.read()
    if len(raw) > 8 * 1024 * 1024:
        raise HTTPException(413, "File troppo grande (massimo 8 MB).")
    parsed = csv_parser.parse_catalog_csv(raw.decode("utf-8-sig", errors="replace"))
    if parsed["errors"]:
        raise HTTPException(400, {"message": "Catalogo non importato: correggi il file.", "errors": parsed["errors"][:30]})
    async with AUCTION_LOCK:
        if auction.mode != "idle":
            raise HTTPException(400, "Attendi la conclusione dell'asta in corso prima di aggiornare il catalogo.")
        try:
            db.replace_catalog(parsed["players"])
        except ValueError as exc:
            raise HTTPException(400, str(exc)) from exc
    storage.salva("catalogo")
    if auto_random_enabled():
        await _arm_auto_random()
    await broadcast_state()
    return {"ok": True, "players_loaded": len(parsed["players"]),
            "out_of_league": sum(p["out_of_league"] for p in parsed["players"]),
            "quotations_loaded": sum("quotazione_mantra" in p["stats"] for p in parsed["players"])}


@app.post("/api/admin/upload/statistics")
async def upload_statistics(file: UploadFile = File(...), _: dict = Depends(require_superadmin)):
    raw = await file.read()
    if len(raw) > 12 * 1024 * 1024:
        raise HTTPException(413, "File troppo grande (massimo 12 MB).")
    name = (file.filename or "").lower()
    try:
        if name.endswith((".xlsx", ".xlsm")):
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            parts = []
            for ws in wb.worksheets:
                rows = [["" if v is None else str(v) for v in row] for row in ws.iter_rows(values_only=True)]
                part = csv_parser.parse_stats_rows(rows)
                # Nei file Strategia/FantaLab ogni ruolo e' su un foglio diverso.
                # Fogli di servizio senza statistiche vengono semplicemente ignorati.
                if part.get("records"):
                    parts.append(part)
            parsed = csv_parser.merge_stats_results(parts)
        else:
            parsed = csv_parser.parse_stats_csv(raw.decode("utf-8-sig", errors="replace"))
    except Exception as exc:
        raise HTTPException(400, f"Impossibile leggere il file statistiche: {exc}") from exc
    if parsed["errors"]:
        raise HTTPException(400, {"message": "Statistiche non importate.", "errors": parsed["errors"][:30]})
    summary = db.update_player_stats(parsed["records"], parsed.get("labels") or {})
    storage.salva("statistiche")
    return {"ok": True, **summary, "fields": list((parsed.get("labels") or {}).values())}


@app.post("/api/admin/upload/rosters")
async def upload_rosters(file: UploadFile = File(...), _: dict = Depends(require_superadmin)):
    if db.auction_mode() != "repair":
        raise HTTPException(400, "Nell'asta da zero non devi caricare un file Rose.")
    raw = await file.read()
    if len(raw) > 8 * 1024 * 1024:
        raise HTTPException(413, "File troppo grande (massimo 8 MB).")
    parsed = csv_parser.parse_rosters_csv(raw.decode("utf-8-sig", errors="replace"))
    if parsed["errors"]:
        raise HTTPException(400, {"message": "Rose non importate: correggi il file.", "errors": parsed["errors"][:30]})
    try:
        budget_summary = db.replace_initial_rosters(list(parsed["teams"].keys()), parsed["assignments"])
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    _stop_timer()
    _cancel_tocca_task()
    _cancel_auto_random()
    db.set_setting("auto_random", "0")
    db.set_setting("auction_started", "0")
    db.set_setting("auction_session_start_event_id", "0")
    auction.reset(persist=True)
    storage.salva("caricamento rose")
    await broadcast_state()
    return {
        "ok": True,
        "teams_found": len(parsed["teams"]),
        "assignments_loaded": len(parsed["assignments"]),
        **budget_summary,
    }


class EnterManagerBody(BaseModel):
    team: str


@app.post("/api/admin/enter-manager")
def enter_manager(body: EnterManagerBody, _: dict = Depends(require_superadmin)):
    team = db.get_team(body.team.strip())
    if not team or not team.get("is_admin"):
        raise HTTPException(400, "Questa squadra non e' configurata come gestore dell'asta.")
    return {
        "token": _new_session("team", team["name"]),
        "team": team["name"],
        "is_manager": True,
        "pin_must_change": bool(team.get("pin_must_change")),
    }


class ResetPinBody(BaseModel):
    team: str


@app.post("/api/admin/teams/reset-pin")
async def admin_reset_pin(body: ResetPinBody, _: dict = Depends(require_superadmin)):
    team = body.team.strip()
    if not db.get_team(team):
        raise HTTPException(404, "Squadra non trovata.")
    temporary = str(secrets.randbelow(9000) + 1000)
    try:
        db.set_team_pin(team, temporary, must_change=True)
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    _invalidate_team_sessions(team)
    await broadcast_state()
    return {"ok": True, "team": team, "temporary_pin": temporary}


class ResetBody(BaseModel):
    keep_catalog: bool = False


@app.post("/api/admin/reset")
async def reset_all(body: ResetBody, session: dict = Depends(require_superadmin)):
    global _timer_task
    _stop_timer()
    _cancel_tocca_task()
    _cancel_auto_random()
    keep_catalog = bool(body.keep_catalog)
    db.reset_all(keep_catalog)
    auction.reset(persist=False)

    # Il reset totale e' una tabula rasa anche per salvataggi e sessioni utente.
    # Manteniamo soltanto la sessione Super Admin che ha eseguito il reset.
    if not keep_catalog:
        storage.pulisci_salvataggi()
        try:
            os.remove(SIMULATION_PATH)
        except FileNotFoundError:
            pass
        for tok in list(SESSIONS):
            if tok != session.get("token"):
                SESSIONS.pop(tok, None)
        _save_sessions()
        manager.last_presence.clear()

    await broadcast_state()
    return {"ok": True, **db.count_rows(), "backups": len(storage.elenco_backup())}


# ========= BACKUP / EXPORT =========
@app.get("/api/admin/export/rose")
def export_rose(_: dict = Depends(require_superadmin)):
    today = datetime.datetime.now().strftime("%Y-%m-%d")
    return Response(
        content=storage.genera_rose_csv(), media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="rose_{today}.csv"'},
    )


@app.get("/api/admin/export/residui")
def export_residui(_: dict = Depends(require_superadmin)):
    today = datetime.datetime.now().strftime("%Y-%m-%d")
    return Response(
        content=storage.genera_residui_csv(), media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="accessi_residui_{today}.csv"'},
    )


@app.post("/api/admin/backup")
def save_backup(_: dict = Depends(require_superadmin)):
    if not storage.salva("manuale"):
        raise HTTPException(500, "Backup non riuscito.")
    return {"ok": True}


@app.get("/api/admin/backup")
def list_backups(_: dict = Depends(require_superadmin)):
    return storage.elenco_backup()


class RestoreBody(BaseModel):
    tag: str


@app.post("/api/admin/backup/restore")
async def restore_backup(body: RestoreBody, _: dict = Depends(require_superadmin)):
    _stop_timer()
    try:
        result = storage.ripristina(body.tag)
    except FileNotFoundError as exc:
        raise HTTPException(404, str(exc)) from exc
    auction.load_persisted()
    if auction.mode in ("bidding", "tiebreak") and not auction.paused:
        _start_timer()
    await broadcast_state()
    return {"ok": True, **result}


# ========= DATI PARTECIPANTE =========
@app.get("/api/state/version")
def get_state_version(response: Response, _: dict = Depends(require_team)):
    _no_store(response)
    return {"version": STATE_VERSION}


@app.get("/api/state")
def get_state(response: Response, session: dict = Depends(require_team)):
    _no_store(response)
    return build_state(session["team"])


@app.get("/api/roster/me")
def my_roster(session: dict = Depends(require_team)):
    return db.get_roster(session["team"])


@app.get("/api/player/{pid}")
def player_detail(pid: int, session: dict = Depends(require_team)):
    _ensure_pin_changed(session)
    p = db.get_player(pid)
    if not p:
        raise HTTPException(404, "Calciatore non trovato.")
    return {"player": p, "stat_labels": db.get_stats_labels()}


@app.get("/api/history")
def history(_: dict = Depends(require_team)):
    return db.get_auction_history(120)


@app.get("/api/rosters")
def rosters(session: dict = Depends(require_team)):
    _ensure_pin_changed(session)
    return {"teams": db.get_public_teams(), "rosters": db.get_all_rosters(), "max_roster": db.max_roster()}


@app.get("/api/free-agents")
def free_agents(session: dict = Depends(require_team)):
    _ensure_pin_changed(session)
    return db.get_free_agents(session["team"])


class FirstPinBody(BaseModel):
    new_pin: str
    team_name: Optional[str] = None


@app.post("/api/account/pin/first")
async def first_pin_change(body: FirstPinBody, session: dict = Depends(require_team)):
    team = db.get_team(session["team"])
    if not team or not team.get("pin_must_change"):
        raise HTTPException(400, "Il cambio PIN iniziale non e' richiesto.")
    old_name = session["team"]
    try:
        updated = db.complete_first_setup(session["team"], body.new_pin, body.team_name)
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    new_name = updated["name"]
    if new_name != old_name:
        for tok, info in list(SESSIONS.items()):
            if info.get("role") == "team" and info.get("team") == old_name:
                if tok == session["token"]:
                    info["team"] = new_name
                else:
                    SESSIONS.pop(tok, None)
        for info in manager.connections.values():
            if info.get("team") == old_name:
                info["team"] = new_name
        if old_name in manager.last_presence:
            manager.last_presence[new_name] = manager.last_presence.pop(old_name)
        _save_sessions()
    else:
        _invalidate_team_sessions(old_name, keep_token=session["token"])
    await broadcast_state()
    return {"ok": True, "team": new_name}


class ChangePinBody(BaseModel):
    current_pin: str
    new_pin: str


@app.post("/api/account/pin/change")
async def change_pin(body: ChangePinBody, session: dict = Depends(require_team)):
    if not db.verify_team_pin(session["team"], body.current_pin):
        raise HTTPException(400, "PIN attuale non corretto.")
    if body.current_pin.strip() == body.new_pin.strip():
        raise HTTPException(400, "Scegli un PIN diverso da quello attuale.")
    try:
        db.set_team_pin(session["team"], body.new_pin, must_change=False)
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    _invalidate_team_sessions(session["team"], keep_token=session["token"])
    return {"ok": True}


@app.get("/api/spectator/state/version")
def spectator_state_version(response: Response, _: dict = Depends(require_spectator)):
    _no_store(response)
    return {"version": STATE_VERSION}


@app.get("/api/spectator/state")
def spectator_state(response: Response, _: dict = Depends(require_spectator)):
    _no_store(response)
    return build_state(None)


@app.get("/api/spectator/player/{pid}")
def spectator_player_detail(pid: int, _: dict = Depends(require_spectator)):
    p = db.get_player(pid)
    if not p:
        raise HTTPException(404, "Calciatore non trovato.")
    return {"player": p, "stat_labels": db.get_stats_labels()}


@app.get("/api/spectator/history")
def spectator_history(_: dict = Depends(require_spectator)):
    return db.get_auction_history(120)


@app.get("/api/spectator/rosters")
def spectator_rosters(_: dict = Depends(require_spectator)):
    return {"teams": db.get_public_teams(), "rosters": db.get_all_rosters(), "max_roster": db.max_roster()}


@app.get("/api/spectator/free-agents")
def spectator_free_agents(_: dict = Depends(require_spectator)):
    return db.get_free_agents(None)


class ChatMessageBody(BaseModel):
    message: str = Field(min_length=1, max_length=500)


@app.get("/api/chat")
def chat_history(after_id: int = 0, limit: int = 120, session: dict = Depends(require_team)):
    _ensure_pin_changed(session)
    return {"messages": db.get_chat_messages(limit=limit, after_id=after_id)}


@app.post("/api/chat")
async def chat_send(body: ChatMessageBody, session: dict = Depends(require_team)):
    _ensure_pin_changed(session)
    try:
        message = db.add_chat_message(session["team"], body.message)
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    await manager.broadcast_team_event("chat_message", message)
    return {"ok": True, "message": message}


@app.delete("/api/chat/{message_id}")
async def chat_delete(message_id: int, _: dict = Depends(require_auction_manager)):
    deleted = db.delete_chat_message(message_id)
    if not deleted:
        raise HTTPException(404, "Messaggio non trovato.")
    await manager.broadcast_team_event("chat_deleted", {"id": int(message_id)})
    return {"ok": True, "id": int(message_id)}


class PushSubscriptionBody(BaseModel):
    endpoint: str
    keys: dict[str, str]


@app.get("/api/push/public-key")
def push_public_key(_: dict = Depends(require_team)):
    return {"public_key": push_service.public_key()}


@app.post("/api/push/subscribe")
def push_subscribe(body: PushSubscriptionBody, session: dict = Depends(require_team)):
    _ensure_pin_changed(session)
    try:
        db.upsert_push_subscription(session["team"], body.model_dump())
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    return {"ok": True}


@app.post("/api/push/unsubscribe")
def push_unsubscribe(body: PushSubscriptionBody, session: dict = Depends(require_team)):
    db.delete_push_subscription(body.endpoint, session["team"])
    return {"ok": True}


class ReadyBody(BaseModel):
    ready: bool = True


@app.post("/api/ready")
async def set_ready(body: ReadyBody, session: dict = Depends(require_team)):
    """Compatibilita' con client precedenti alla v10.

    Da v10 non esiste piu' il toggle Pronto: un Rannatone che ha effettuato
    il login resta in asta fino a quando conclude il mercato.
    """
    team = _ensure_pin_changed(session)
    if team and team.get("market_finished"):
        raise HTTPException(400, "Hai gia' concluso il mercato.")
    db.set_ready(session["team"], True)
    await broadcast_state()
    if auction.mode == "idle" and auto_random_enabled():
        await _arm_auto_random()
    return {"ok": True, "ready": True}


@app.post("/api/market/finish")
async def finish_market(session: dict = Depends(require_team)):
    _ensure_pin_changed(session)
    _ensure_auction_started()
    if auction.mode != "idle":
        raise HTTPException(400, "Puoi concludere il mercato solo tra un'asta e la successiva.")
    summary = db.roster_summary(session["team"])
    if not summary["valid_finish"]:
        raise HTTPException(400, {
            "message": "La rosa non rispetta i requisiti per concludere il mercato.",
            "roster_rules": summary,
        })
    try:
        db.set_market_finished(session["team"], True)
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    # "Ho finito gli acquisti" e disconnessione sono un'unica operazione:
    # invalida tutte le sessioni della squadra. Il client pulisce poi il token locale.
    _invalidate_team_sessions(session["team"])
    await broadcast_state()
    await _arm_auto_random()
    return {"ok": True, "market_finished": True, "logged_out": True, "roster_rules": summary}


# ========= TIMER / ASTA =========
async def _push_new_auction():
    player = db.get_player(auction.current_pid) if auction.current_pid is not None else None
    if not player or not auction.eligible_teams:
        return

    # v13: unico messaggio automatico della chat.
    system_message = db.add_system_chat_message(f"🎲 È iniziata l’asta per {player['name']}")
    if system_message:
        await manager.broadcast_team_event("chat_message", system_message)

    await push_service.send_to_teams(
        auction.eligible_teams,
        title="🎲 Nuova asta Rannatoni",
        body=f"{player['name']} — entra per offrire",
        url="/auction",
        tag=f"asta-{player['pid']}",
    )


async def _push_tiebreak(result: dict):
    player = db.get_player(result.get("pid")) if result.get("pid") is not None else None
    if not player:
        return
    await push_service.send_to_teams(
        result.get("teams") or [],
        title="⚔️ Spareggio Rannatoni",
        body=f"{player['name']} — devi rilanciare",
        url="/auction",
        tag=f"spareggio-{player['pid']}",
    )


def _cancel_auto_random():
    global _auto_random_task, _auto_random_deadline
    task = _auto_random_task
    _auto_random_task = None
    _auto_random_deadline = None
    if task and not task.done() and task is not asyncio.current_task():
        task.cancel()


async def _arm_auto_random():
    """Avvia il conto alla rovescia se c'e' almeno un Rannatone entrato nel mercato."""
    global _auto_random_task, _auto_random_deadline
    _cancel_auto_random()
    if (not auction_started() and not simulation_active()) or not auto_random_enabled() or auction.mode != "idle":
        return False
    ready = db.ready_market_teams()
    if not ready:
        await broadcast_state()
        return False
    if not db.random_player_pool()["ids"]:
        await broadcast_state()
        return False
    _auto_random_deadline = time.time() + db.auto_random_delay()

    async def runner():
        global _auto_random_task, _auto_random_deadline
        try:
            while _auto_random_deadline and time.time() < _auto_random_deadline:
                await asyncio.sleep(0.25)
            async with AUCTION_LOCK:
                # Ricontrolla tutto allo zero: qualcuno potrebbe aver concluso
                # il mercato o l'admin potrebbe aver aperto manualmente.
                if not auto_random_enabled() or auction.mode != "idle":
                    return
                ready_now = db.ready_market_teams()
                if not ready_now:
                    return
                auction.open_random()
                _start_timer()
            await broadcast_state()
            asyncio.create_task(_push_new_auction())
        except asyncio.CancelledError:
            return
        except auction_module.AuctionError as exc:
            print(f"[auto-random] {exc}")
        finally:
            # Un cambio soglia puo' aver gia' avviato un nuovo countdown.
            # Il task cancellato non deve azzerare quello che lo sostituisce.
            if _auto_random_task is asyncio.current_task():
                _auto_random_task = None
                _auto_random_deadline = None
            await broadcast_state()

    _auto_random_task = asyncio.create_task(runner())
    await broadcast_state()
    return True


def _stop_timer():
    global _timer_task
    task = _timer_task
    _timer_task = None
    if task and not task.done() and task is not asyncio.current_task():
        task.cancel()


def _start_timer():
    global _timer_task
    _stop_timer()
    if auction.mode not in ("bidding", "tiebreak") or auction.paused:
        return

    async def waiter():
        try:
            while auction.mode in ("bidding", "tiebreak") and not auction.paused:
                if auction.expired():
                    async with AUCTION_LOCK:
                        if auction.mode in ("bidding", "tiebreak") and auction.expired():
                            await _close_and_broadcast(from_timer=True)
                    return
                await asyncio.sleep(0.5)
        except asyncio.CancelledError:
            return

    _timer_task = asyncio.create_task(waiter())


def _public_tocca_result(result: dict):
    """Nasconde il vincitore finche' la TOCCA non e' terminata."""
    if not result or not result.get("tocca_pending"):
        return result
    public = dict(result)
    public.pop("team", None)
    public.pop("needs_release", None)
    return public


def _cancel_tocca_task():
    global _tocca_task
    task = _tocca_task
    _tocca_task = None
    if task and not task.done() and task is not asyncio.current_task():
        task.cancel()


def _tocca_remaining_seconds(result: dict | None = None):
    r = result or auction.last_result or {}
    if not r.get("tocca") or not r.get("tocca_pending"):
        return 0.0
    started_ms = float(r.get("tocca_reveal_started_at_ms") or 0)
    duration_ms = float(r.get("tocca_reveal_duration_ms") or auction_module.TOCCA_REVEAL_MS)
    if started_ms <= 0:
        return 0.0
    return max(0.0, (started_ms + duration_ms) / 1000 - time.time())


def _schedule_tocca_finalize(result: dict | None = None):
    """Conclude la TOCCA lato server allo scadere del reveal globale."""
    global _tocca_task
    _cancel_tocca_task()
    expected = dict(result or auction.last_result or {})
    reveal_id = expected.get("tocca_reveal_id")
    if auction.mode != "tocca" or not reveal_id:
        return False

    async def runner():
        global _tocca_task
        try:
            remaining = _tocca_remaining_seconds(expected)
            if remaining > 0:
                await asyncio.sleep(remaining)
            async with AUCTION_LOCK:
                current = auction.last_result or {}
                if auction.mode != "tocca" or current.get("tocca_reveal_id") != reveal_id:
                    return
                final = auction.finalize_tocca()
                if not final.get("needs_release"):
                    backup_if_real(f"acquisto {final.get('player_name', '')}")
            # Solo ora il nome del vincitore viene trasmesso a tutti.
            await manager.broadcast_result(final)
            await broadcast_state()
            if not final.get("needs_release"):
                await _arm_auto_random()
        except asyncio.CancelledError:
            return
        except Exception as exc:
            print(f"[tocca] finalizzazione non riuscita: {exc}")
        finally:
            if _tocca_task is asyncio.current_task():
                _tocca_task = None

    _tocca_task = asyncio.create_task(runner())
    return True


async def _close_and_broadcast(from_timer: bool = False):
    result = auction.close()
    tocca_pending = bool(result.get("tocca_pending"))
    if result.get("type") == "tiebreak":
        _start_timer()
    else:
        _stop_timer()
        if not tocca_pending:
            if result.get("type") == "assigned" and not result.get("needs_release"):
                backup_if_real(f"acquisto {result.get('player_name', '')}")
            elif result.get("type") == "no_offers":
                backup_if_real("giocatore passato")
    await manager.broadcast_result(_public_tocca_result(result))
    await broadcast_state()
    if result.get("type") == "tiebreak":
        asyncio.create_task(_push_tiebreak(result))
    elif tocca_pending:
        # Nessun altro passaggio dell'asta puo' partire prima della fine della TOCCA.
        _schedule_tocca_finalize(result)
    elif result.get("type") in ("assigned", "no_offers") and not result.get("needs_release"):
        await _arm_auto_random()
    return result


def _readiness_status():
    active = db.active_market_teams()
    missing = [t["name"] for t in active if not t.get("ready")]
    return len(active), missing


class RandomBody(BaseModel):
    force: bool = False


@app.post("/api/auction/admin/random")
async def open_random(body: RandomBody, _: dict = Depends(require_auction_manager)):
    _ensure_auction_started()
    _cancel_auto_random()
    async with AUCTION_LOCK:
        try:
            pid = auction.open_random()
        except auction_module.AuctionError as exc:
            raise HTTPException(400, str(exc)) from exc
        _start_timer()
    await broadcast_state()
    asyncio.create_task(_push_new_auction())
    return {"ok": True, "pid": pid, "player": db.get_player(pid), "participants": len(auction.eligible_teams)}


class OpenPlayerBody(BaseModel):
    pid: int
    force: bool = False


@app.post("/api/auction/admin/open")
async def open_manual(body: OpenPlayerBody, _: dict = Depends(require_auction_manager)):
    _ensure_auction_started()
    _cancel_auto_random()
    async with AUCTION_LOCK:
        try:
            auction.open_player(body.pid)
        except auction_module.AuctionError as exc:
            raise HTTPException(400, str(exc)) from exc
        _start_timer()
    await broadcast_state()
    asyncio.create_task(_push_new_auction())
    return {"ok": True}


@app.get("/api/auction/admin/search")
def manager_search(q: str = "", _: dict = Depends(require_auction_manager)):
    players = db.search_players(q, limit=30, exclude_assigned=True)
    passed_ids = {p["pid"] for p in db.get_passed_players()}
    for p in players:
        p["passed"] = p["pid"] in passed_ids
    return players


@app.post("/api/auction/admin/close")
async def close_bids(session: dict = Depends(require_auction_manager)):
    async with AUCTION_LOCK:
        if auction.mode not in ("bidding", "tiebreak"):
            raise HTTPException(400, "Nessuna asta aperta.")
        if session["team"] in auction.eligible_teams and session["team"] not in auction.offers:
            raise HTTPException(400, "Prima di aprire le buste devi inviare la tua risposta.")
        # Il gestore puo' forzare l'apertura in qualunque momento: chi manca
        # viene considerato PASSO (o mantiene la base nello spareggio).
        result = await _close_and_broadcast()
    return result


class BidBody(BaseModel):
    amount: int = Field(gt=0)


@app.post("/api/auction/bid")
async def bid(body: BidBody, session: dict = Depends(require_team)):
    _ensure_pin_changed(session)
    result = None
    async with AUCTION_LOCK:
        try:
            auction.bid(session["team"], body.amount)
        except auction_module.AuctionError as exc:
            raise HTTPException(400, str(exc)) from exc
        if auction.all_submitted():
            result = await _close_and_broadcast()
    if result is None:
        await broadcast_state()
    return {"ok": True, "amount": body.amount, "auto_opened": result is not None}


@app.post("/api/auction/pass")
async def pass_bid(session: dict = Depends(require_team)):
    _ensure_pin_changed(session)
    result = None
    async with AUCTION_LOCK:
        try:
            auction.pass_(session["team"])
        except auction_module.AuctionError as exc:
            raise HTTPException(400, str(exc)) from exc
        if auction.all_submitted():
            result = await _close_and_broadcast()
    if result is None:
        await broadcast_state()
    return {"ok": True, "auto_opened": result is not None}


class ReleaseBody(BaseModel):
    pids: list[int]


@app.post("/api/auction/release")
async def release_for_purchase(body: ReleaseBody, session: dict = Depends(require_team)):
    info = auction.last_result.get("needs_release", {}) if auction.last_result else {}
    if auction.mode != "release" or info.get("team") != session["team"]:
        raise HTTPException(403, "Solo il vincitore puo' completare gli svincoli.")
    async with AUCTION_LOCK:
        try:
            result = auction.complete_with_releases(body.pids)
        except auction_module.AuctionError as exc:
            raise HTTPException(400, str(exc)) from exc
        backup_if_real("acquisto con svincoli")
    await manager.broadcast_result(result)
    await broadcast_state()
    await _arm_auto_random()
    return result


class AddTimeBody(BaseModel):
    seconds: int = Field(gt=0, le=300)


@app.post("/api/auction/admin/pause")
async def pause_timer(_: dict = Depends(require_auction_manager)):
    async with AUCTION_LOCK:
        try:
            auction.pause()
        except auction_module.AuctionError as exc:
            raise HTTPException(400, str(exc)) from exc
        _stop_timer()
    await broadcast_state()
    return {"ok": True}


@app.post("/api/auction/admin/resume")
async def resume_timer(_: dict = Depends(require_auction_manager)):
    async with AUCTION_LOCK:
        try:
            auction.resume()
        except auction_module.AuctionError as exc:
            raise HTTPException(400, str(exc)) from exc
        _start_timer()
    await broadcast_state()
    return {"ok": True}


@app.post("/api/auction/admin/add-time")
async def add_time(body: AddTimeBody, _: dict = Depends(require_auction_manager)):
    async with AUCTION_LOCK:
        try:
            auction.add_time(body.seconds)
        except auction_module.AuctionError as exc:
            raise HTTPException(400, str(exc)) from exc
    await broadcast_state()
    return {"ok": True}


@app.post("/api/auction/admin/cancel")
async def cancel_auction(_: dict = Depends(require_auction_manager)):
    _cancel_tocca_task()
    _cancel_auto_random()
    async with AUCTION_LOCK:
        try:
            auction.cancel()
        except auction_module.AuctionError as exc:
            raise HTTPException(400, str(exc)) from exc
        _stop_timer()
    await broadcast_state()
    return {"ok": True}


class AutoRandomBody(BaseModel):
    enabled: bool


@app.post("/api/auction/admin/auto-random")
async def set_auto_random(body: AutoRandomBody, _: dict = Depends(require_auction_manager)):
    _ensure_auction_started()
    db.set_setting("auto_random", "1" if body.enabled else "0")
    if body.enabled:
        await _arm_auto_random()
    else:
        _cancel_auto_random()
        await broadcast_state()
    return {"ok": True, "enabled": body.enabled}


@app.post("/api/auction/admin/auto-random/skip")
async def skip_next_auto_random(_: dict = Depends(require_auction_manager)):
    _cancel_auto_random()
    await broadcast_state()
    return {"ok": True, "enabled": auto_random_enabled()}


class ReactivateTeamBody(BaseModel):
    team: str


@app.post("/api/auction/admin/reactivate-team")
async def reactivate_team(body: ReactivateTeamBody, _: dict = Depends(require_auction_manager)):
    if auction.mode != "idle":
        raise HTTPException(400, "Riattiva una squadra tra un'asta e la successiva.")
    team_name = body.team.strip()
    try:
        db.set_market_finished(team_name, False)
        # Se il Rannatone e' gia' autenticato in sola consultazione, la riattivazione
        # lo rimette subito in asta. Altrimenti entrera' automaticamente al login.
        has_live_session = any(
            sess.get("role") == "team" and sess.get("team") == team_name
            and float(sess.get("expires_at", 0)) > time.time()
            for sess in SESSIONS.values()
        )
        if has_live_session:
            db.set_ready(team_name, True)
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    await broadcast_state()
    if auction.mode == "idle" and auto_random_enabled():
        await _arm_auto_random()
    return {"ok": True}


@app.post("/api/auction/admin/reset-ready")
async def reset_ready(_: dict = Depends(require_auction_manager)):
    # Endpoint legacy: da v10 lo stato "Pronto" non e' piu' modificabile manualmente.
    raise HTTPException(410, "Da v10 i Rannatoni entrano automaticamente nel mercato al login.")


@app.get("/api/auction/admin/passed")
def passed_players(_: dict = Depends(require_auction_manager)):
    return db.get_passed_players()


class ReopenPassedBody(BaseModel):
    pid: Optional[int] = None


@app.post("/api/auction/admin/reopen-passed")
async def reopen_passed(body: ReopenPassedBody, _: dict = Depends(require_auction_manager)):
    db.clear_passed(body.pid)
    await broadcast_state()
    return {"ok": True}


@app.post("/api/auction/admin/undo-last")
async def undo_last(_: dict = Depends(require_auction_manager)):
    _cancel_auto_random()
    if auction.mode != "idle":
        raise HTTPException(400, "Puoi annullare l'ultima assegnazione solo quando non c'e' un'asta in corso.")
    async with AUCTION_LOCK:
        try:
            result = db.undo_last_assignment()
        except ValueError as exc:
            raise HTTPException(400, str(exc)) from exc
        auction.last_result = None
        auction._persist()
        backup_if_real("annulla ultima assegnazione")
    await broadcast_state()
    return {"ok": True, **result}


# ========= MODALITA' SIMULAZIONE =========
@app.post("/api/auction/admin/simulation/start")
async def start_simulation(_: dict = Depends(require_auction_manager)):
    _cancel_auto_random()
    if simulation_active():
        raise HTTPException(400, "La simulazione e' gia' attiva.")
    if auction.mode != "idle":
        raise HTTPException(400, "Termina o annulla l'asta corrente prima di avviare la simulazione.")
    async with AUCTION_LOCK:
        db.backup_database(SIMULATION_PATH)
        db.set_setting("simulation_active", "1")
        auction.last_result = None
        auction._persist()
    await broadcast_state()
    return {"ok": True}


@app.post("/api/auction/admin/simulation/end")
async def end_simulation(_: dict = Depends(require_auction_manager)):
    _cancel_auto_random()
    if not simulation_active() or not os.path.exists(SIMULATION_PATH):
        raise HTTPException(400, "Nessuna simulazione da terminare.")
    async with AUCTION_LOCK:
        _stop_timer()
        db.restore_database(SIMULATION_PATH)
        auction.load_persisted()
        try:
            os.remove(SIMULATION_PATH)
        except FileNotFoundError:
            pass
        if auction.mode in ("bidding", "tiebreak") and not auction.paused:
            _start_timer()
    await broadcast_state()
    return {"ok": True}


# ========= WEBSOCKET =========
@app.websocket("/ws")
async def websocket_endpoint(ws: WebSocket, token: str):
    session = SESSIONS.get(token)
    if not session or float(session.get("expires_at", 0)) <= time.time():
        await ws.close(code=4401)
        return
    await manager.connect(ws, session)
    try:
        await manager.send_state(ws)
        await broadcast_state()
        while True:
            raw = await ws.receive_text()
            try:
                msg = json.loads(raw)
            except Exception:
                msg = {"type": "ping", "visibility": "visible"}
            if msg.get("type") in ("presence", "ping"):
                changed = manager.touch(ws, msg.get("visibility", "visible"))
                if msg.get("type") == "ping":
                    # Risposta immediata: la PWA usa il pong per distinguere un
                    # WebSocket davvero vivo da un socket mobile rimasto OPEN
                    # ma sospeso dal sistema operativo.
                    try:
                        await manager._send_with_timeout(ws, {
                            "type": "pong",
                            "server_now_ms": int(time.time() * 1000),
                        }, timeout=2.0)
                    except Exception:
                        break
                if changed:
                    await broadcast_state()
    except WebSocketDisconnect:
        pass
    finally:
        manager.disconnect(ws)
        await broadcast_state()


@app.on_event("startup")
async def startup_resume_timer():
    global _presence_task
    if auction.mode in ("bidding", "tiebreak") and not auction.paused:
        _start_timer()
    elif auction.mode == "tocca":
        # Se Railway riavvia il container durante il reveal, riprende dalla
        # scadenza persistita e non perde l'assegnazione.
        _schedule_tocca_finalize(auction.last_result)

    async def presence_loop():
        try:
            while True:
                await asyncio.sleep(15)
                await broadcast_state()
        except asyncio.CancelledError:
            return

    _presence_task = asyncio.create_task(presence_loop())


@app.on_event("shutdown")
async def shutdown_tasks():
    global _presence_task
    _stop_timer()
    _cancel_tocca_task()
    _cancel_auto_random()
    if _presence_task and not _presence_task.done():
        _presence_task.cancel()
    _presence_task = None


# ========= FRONTEND / PWA =========
app.mount("/static", StaticFiles(directory=FRONTEND_DIR), name="static")
NO_CACHE = {
    "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
    "Pragma": "no-cache",
    "Expires": "0",
}


@app.get("/health")
def health():
    return {"ok": True, "service": "rannatoni"}


@app.get("/spectator")
def serve_spectator():
    return FileResponse(os.path.join(FRONTEND_DIR, "spectator.html"), headers=NO_CACHE)


@app.get("/")
def serve_index():
    return FileResponse(os.path.join(FRONTEND_DIR, "index.html"), headers=NO_CACHE)


@app.get("/admin")
def serve_admin():
    return FileResponse(os.path.join(FRONTEND_DIR, "admin.html"), headers=NO_CACHE)


@app.get("/auction")
def serve_auction():
    return FileResponse(os.path.join(FRONTEND_DIR, "auction.html"), headers=NO_CACHE)


@app.get("/manifest.webmanifest")
def manifest():
    return FileResponse(os.path.join(FRONTEND_DIR, "manifest.webmanifest"), media_type="application/manifest+json")


@app.get("/sw.js")
def service_worker():
    return FileResponse(
        os.path.join(FRONTEND_DIR, "sw.js"), media_type="application/javascript",
        headers={"Cache-Control": "no-cache"},
    )
